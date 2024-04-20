from my_functions import get_used_file, show_alert, read_config, remove_folder_contents
import openpyxl
import os
import shutil
import sys

from datetime import datetime

config_values = read_config("config.json") if sys.platform.startswith(
    'win') else read_config("config_mac.json")

database_file_path = config_values.get('database_file_path', '')
usage_addition_file_path = config_values.get('usage_addition_file_path', '')

done_folder_path = config_values.get('done_folder_path', '')

usage_worksheet = config_values.get('usage_worksheet', '')
usage_workdatasheet = config_values.get('usage_workdatasheet', '')

addition_start_row = int(config_values.get('addition_start_row', ''))
database_start_row = int(config_values.get('database_start_row', ''))

temp = '.\\temp' if sys.platform.startswith('win') else './temp'



def find_missing_records(database_path, usage_path):
    try:
        database_wb = openpyxl.load_workbook(database_path)
        todo_file_path, todo_file_name = get_used_file(usage_path)
        usage_wb = openpyxl.load_workbook(todo_file_path)

        database_sheet = database_wb[usage_workdatasheet]
        usage_sheet = usage_wb[usage_worksheet]
        missing_records = []

        for usage_row in usage_sheet.iter_rows(min_row=addition_start_row, values_only=True):
            asset_name, _ = usage_row[1], usage_row[3]

            asset_found = False
            for database_row in database_sheet.iter_rows(min_row=database_start_row, max_row=database_sheet.max_row, values_only=True):
                asset_number = database_row[1]

                if asset_number == asset_name:
                    asset_found = True
                    break

            if not asset_found and asset_name is not None:
                missing_records.append(asset_name)

        print(
            f"Tych elementow nie ma w pliku magazynowym: {missing_records} a sa dostepne w pliku: {todo_file_name}")
    except KeyError as e:
        print(f"Error: {e}")
        show_alert(
            "Skoroszyt Excela nie nazywa sie 'Materialliste'", f"{str(e)}")
    except FileNotFoundError as e:
        show_alert("Brakuje pliku", f"{str(e)}")
    except PermissionError as e:
        show_alert("Excel jest wciaz otwarty", f"{e}")


def update_quantities(database_path, usage_path, done_folder_path, temp):
    try:
        database_wb = openpyxl.load_workbook(database_path)
        todo_file_path, todo_file_name = get_used_file(usage_path)
        shutil.move(todo_file_path, temp)
        todo_file_path_temp = os.path.join(temp, todo_file_name)
        usage_wb = openpyxl.load_workbook(todo_file_path_temp)
        # usage_wb = openpyxl.load_workbook(usage_path)

        database_sheet = database_wb[usage_workdatasheet]
        usage_sheet = usage_wb[usage_worksheet]
        # print(database_sheet)
        count = 0
        negative_record = []
        for row in usage_sheet.iter_rows(min_row=addition_start_row, values_only=True):
            asset_name, used_quantity = row[1], row[3]
            # print(asset_name, used_quantity)
            # print(row)

            for index, database_row in enumerate(database_sheet.iter_rows(min_row=database_start_row, max_row=database_sheet.max_row, values_only=True), start=2):
                asset_number, current_quantity = database_row[1], database_row[3]

                if asset_number == asset_name:

                    if current_quantity is not None and used_quantity is not None:
                        # new_quantity = max(current_quantity - used_quantity, 0)
                        new_quantity = current_quantity + used_quantity
                        # Here we could modify it to: int(used_quantity.rstrip("m"))
                        modified_cell = database_sheet.cell(
                            row=index, column=4)
                        modified_cell.value = new_quantity
                        count += 1
                        print(
                            f"Wiersz zmieniony: {index}, Numer czesci: {asset_number}")
                        if new_quantity < 0:
                            print(
                                f"!!!!! Elemnt z wiersza: {index} ma UJEMNA wartosc: {new_quantity}. Numer czesci: {asset_number} !!!")
                            negative_record.append(asset_number)

        print(f"Liczba elementow ktore zostaly zedytowane: {count}")
        print(
            f"Te elementy na magazynie maja ujemna wartosc: {negative_record}")
        try:
            print(f"Porces zapisu pliku rozpoczety: {database_path}")
            
            database_wb.close()
            database_wb.save(database_path)
            print(f"Proces zapisu zakonczony: {database_path}")
            new_name = todo_file_name.split(".")[0] + datetime.now().strftime(
                "%Y-%m-%d %H-%M-%S") + "." + todo_file_name.split(".")[1]
            done_file_path = os.path.join(done_folder_path, new_name)
           
            
            shutil.copy(todo_file_path_temp, done_file_path)
            usage_wb.close()
            print(
                f"Plik z materialami obrobiony i przeniesiony do folderu: {done_file_path}")
        except PermissionError as e:
            print(f"Error: {e}")
            show_alert("Ograniczony dostep do pliku",
                       f"Baza danych nie zostala zapisana, gdyz dostep byl ograniczony. Prawdopodobnie Excel z baza danych jest otwarty. {str(e)}. Plik NIE zostal zapisany. Zamknij go i odpal skrypt ponownie")
        except TypeError as e:
            show_alert(f"W pliku istnieja metry i nie policzylem: {e}")
    except shutil.Error as e:
        print(f"Error: {e}")
        show_alert("W Folderze istnieje juz plik o tej nazwie",
                   f"Baza danych Zostala zaktualizowana, wiec musisz tylko przeniesc plik do folderu '{done_file_path}': {str(e)}")
    except KeyError as e:
        print(f"Error: {e}")
        show_alert(
            "Skoroszyt Excela nie nazywa sie 'Materialliste'", f"{str(e)}")
    except TypeError as e:
        show_alert("W excelu nie bylo podanych numerow",
                   f"W pliku istnieja metry i nie policzylem: {e}")
    except PermissionError as e:
            print(f"Error: {e}")
            show_alert("Ograniczony dostep do pliku",
                       f"Baza danych nie zostala zapisana, gdyz dostep byl ograniczony. Prawdopodobnie Excel z baza danych jest otwarty. {str(e)}. Plik NIE zostal zapisany. Zamknij go i odpal skrypt ponownie")


if __name__ == "__main__":
    print("Script starting")
    remove_folder_contents(temp)
    find_missing_records(database_file_path, usage_addition_file_path)
    update_quantities(database_file_path,
                      usage_addition_file_path, done_folder_path)
    print(50*"-")
    # input("Press Enter to exit...")
